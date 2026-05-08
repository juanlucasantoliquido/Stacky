"""
report_output_verifier.py — Report Output Structure Test.

Verifies generated reports (Excel, CSV) have correct columns, non-empty data,
and no error cells.

Uso:
    from report_output_verifier import ReportOutputVerifier
    verifier = ReportOutputVerifier(config)
    cases = verifier.verify(ticket_folder, config)
"""

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

logger = logging.getLogger("stacky.report_output")


@dataclass
class ReportTestCase:
    name: str
    passed: bool
    evidence: str = ""


class ReportOutputVerifier:
    def __init__(self, config: Optional[dict] = None,
                 batch_executor=None):
        self.config = config or {}
        self.batch_executor = batch_executor

    def verify(self, ticket_folder: str, config: Optional[dict] = None) -> list[ReportTestCase]:
        cfg = config or self.config
        if not self._generates_report(ticket_folder):
            return []

        expected_spec = self._extract_report_spec(ticket_folder)

        # Execute process that generates the report
        if self.batch_executor:
            self.batch_executor.run_minimal(ticket_folder)

        report_file = self._find_generated_report(ticket_folder, cfg)
        if not report_file:
            return [ReportTestCase(
                name="Report file generated",
                passed=False,
                evidence="Report file not found"
            )]

        cases = []

        if report_file.suffix == ".xlsx":
            cases.extend(self._verify_xlsx(report_file, expected_spec))
        elif report_file.suffix == ".csv":
            cases.extend(self._verify_csv(report_file, expected_spec))
        else:
            cases.append(ReportTestCase(
                name="Report file exists",
                passed=True,
                evidence=f"File: {report_file.name}, size={report_file.stat().st_size}"
            ))

        return cases

    def _verify_xlsx(self, report_file: Path, spec: dict) -> list[ReportTestCase]:
        cases = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(report_file)
            ws = wb.active

            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

            # Check expected columns
            for expected_col in spec.get("columns", []):
                cases.append(ReportTestCase(
                    name=f"Column '{expected_col}' present",
                    passed=(expected_col in headers),
                    evidence=f"Headers: {headers[:10]}"
                ))

            # Check not empty
            cases.append(ReportTestCase(
                name="Report has data rows",
                passed=(ws.max_row > 1),
                evidence=f"Rows: {ws.max_row}"
            ))

            # Check no error cells
            error_cells = []
            for r in range(2, min(ws.max_row + 1, 100)):
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(r, c).value or "")
                    if val.startswith("#") and any(k in val for k in ["REF", "ERROR", "DIV", "NAME"]):
                        error_cells.append(f"({r},{c})")

            cases.append(ReportTestCase(
                name="No error cells (#REF!, #ERROR#)",
                passed=not error_cells,
                evidence=f"Error cells: {error_cells[:5]}" if error_cells else "Clean"
            ))

        except ImportError:
            cases.append(ReportTestCase(
                name="openpyxl available", passed=False,
                evidence="openpyxl not installed"
            ))
        except Exception as e:
            cases.append(ReportTestCase(
                name="XLSX verification", passed=False,
                evidence=str(e)[:200]
            ))

        return cases

    def _verify_csv(self, report_file: Path, spec: dict) -> list[ReportTestCase]:
        cases = []
        try:
            content = report_file.read_text(encoding="utf-8", errors="replace")
            lines = content.strip().split("\n")

            cases.append(ReportTestCase(
                name="CSV has data rows",
                passed=(len(lines) > 1),
                evidence=f"Lines: {len(lines)}"
            ))

        except Exception as e:
            cases.append(ReportTestCase(
                name="CSV verification", passed=False,
                evidence=str(e)[:200]
            ))

        return cases

    def _generates_report(self, ticket_folder: str) -> bool:
        folder = Path(ticket_folder)
        for fname in ["TAREAS_DESARROLLO.md", "ARQUITECTURA_SOLUCION.md"]:
            p = folder / fname
            if p.exists():
                content = p.read_text(encoding="utf-8", errors="replace").lower()
                if any(k in content for k in [
                    "reporte", "report", "excel", "xlsx", "csv",
                    "listado", "exportar"
                ]):
                    return True
        return False

    def _extract_report_spec(self, ticket_folder: str) -> dict:
        folder = Path(ticket_folder)
        columns = []
        for fname in ["TAREAS_DESARROLLO.md", "ARQUITECTURA_SOLUCION.md"]:
            p = folder / fname
            if p.exists():
                content = p.read_text(encoding="utf-8", errors="replace")
                for m in re.finditer(r"columna[s]?\s*:\s*(.+?)[\n\.]", content, re.IGNORECASE):
                    cols = [c.strip() for c in m.group(1).split(",")]
                    columns.extend(cols)
        return {"columns": columns}

    def _find_generated_report(self, ticket_folder: str, config: dict) -> Optional[Path]:
        output_dir = config.get("report_output_dir", ticket_folder)
        folder = Path(output_dir)
        for ext in ["*.xlsx", "*.csv", "*.xls"]:
            files = list(folder.glob(ext))
            if files:
                return sorted(files, key=lambda f: f.stat().st_mtime, reverse=True)[0]
        return None
